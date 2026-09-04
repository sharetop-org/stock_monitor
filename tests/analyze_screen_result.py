#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""统计「全市场回测扫描」导出 Excel 中『利润 / 总收益率 / 复合年化 均>0』的股票占比。

读取 screen-backtest --excel 导出的结果文件(data/screen_backtest_result.xlsx),
第一行为标题, 列含: symbol, name, list_year, 利润(万), 总收益%, 复合年化%, 胜率%。

对每一行判断:
    利润(万) > 0  AND  总收益% > 0  AND  复合年化% > 0
统计满足条件的股票数占总股票数量的百分比, 并打印。

用法:
    python tests/analyze_screen_result.py [文件路径]
    缺省文件名: data/screen_backtest_result.xlsx

说明: 用 openpyxl.load_workbook 直接读取(不依赖 pandas.read_excel),
      可兼容较旧的 openpyxl(如 3.0.x)。
"""
from __future__ import annotations

import os
import sys
from typing import List, Optional, Tuple

from openpyxl import load_workbook

DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "screen_backtest_result.xlsx")

# 三盈口径需要匹配的标题列名
COLS = ["利润(万)", "总收益%", "复合年化%"]


def _to_number(val) -> Optional[float]:
    """把单元格值转 float; 非数字/空返回 None。"""
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in {"-", "—", "--", "nan", "None"}:
        return None
    # 处理可能的百分号/逗号等
    s = s.replace("%", "").replace(",", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def load_rows(path: str) -> Tuple[List[str], List[List[object]]]:
    """读取第一个工作表, 返回 (标题行, 数据行列表)。"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在: {path}（请先跑 `screen-backtest --excel ...` 生成）")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(c) for c in next(it)]
    except StopIteration:
        wb.close()
        raise ValueError(f"文件为空: {path}")
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def win_ratio(header: List[str], rows: List[List[object]]) -> Tuple[int, int, float]:
    """返回 (三盈股票数, 总股票数, 三盈占比%)。

    缺失/不可解析的单元格不计为"三盈"但计入总数; 行数比标题略短按 None 处理。
    """
    idx = {}
    for c in COLS:
        if c in header:
            idx[c] = header.index(c)
    if len(idx) < len(COLS):
        missing = [c for c in COLS if c not in header]
        raise ValueError(f"缺少必需列: {missing}(实际标题: {header})")

    total = len(rows)
    ok = 0
    for r in rows:
        vals = []
        valid = True
        for c in COLS:
            v = _to_number(r[idx[c]] if idx[c] < len(r) else None)
            if v is None:
                valid = False
                break
            vals.append(v)
        if valid and vals[0] > 0 and vals[1] > 0 and vals[2] > 0:
            ok += 1
    pct = (ok / total * 100) if total else 0.0
    return (ok, total, pct)


def main(argv: Optional[List[str]] = None) -> None:
    args = argv if argv is not None else sys.argv
    path = args[1] if len(args) > 1 else DEFAULT_FILE
    header, rows = load_rows(path)
    ok, total, pct = win_ratio(header, rows)
    print(f"文件                     : {path}")
    print(f"总股票数量              : {total}")
    print(f"三盈(利润>0 & 总收益>0 & 复合年化>0) 只数: {ok}")
    print(f"三盈占比                 : {pct:.2f}%")


if __name__ == "__main__":
    main()